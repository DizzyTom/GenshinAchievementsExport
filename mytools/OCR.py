from openpyxl.worksheet.worksheet import Worksheet
from .VRBS import vrbs
from .RECTS import *
from paddleocr import PaddleOCR
import cv2
import numpy as np
from .TEXT import Find, Find2
from .ACTION import rscroll, screen_shot
import os
import openpyxl
ocr= PaddleOCR(lang='ch')

def Image2Result(image):   
    result= ocr.ocr(image)
    texts=[]
    for line in result:
        text=line[1][0]
        texts.append(text)
    return texts
def get_left_infos(image):
    image=cv2.cvtColor(np.asarray(image),cv2.COLOR_RGB2BGR)
    image_canny=cv2.Canny(image,50,100)
    contours=cv2.findContours(image_canny,cv2.CHAIN_APPROX_SIMPLE,cv2.RETR_CCOMP)[0]
    rects=[cv2.cv2.boundingRect(contour) for contour in contours]
    left_rects= filter_rects(rects,vrbs.left_area,vrbs.LNW, vrbs.LXW ,vrbs.LNH , vrbs.LXH)
    left_rects= short_rects(left_rects)
    left_rects.sort(key=lambda x:x[1])
    texts=[]
    for rect in left_rects:
        left=rect[0]
        up=rect[1]
        right=rect[0]+rect[2]
        down=rect[1]+rect[3]
        left=left+int( vrbs.LRT/image.shape[1]*(right-left))
        text= Find(Image2Result( image[up:down,left:right])[0])
        texts.append(text)
    return texts,left_rects

def get_right_images(name):
    last_img=[]
    while True:
        image=screen_shot(name)        
        image=cv2.cvtColor(np.asarray(image),cv2.COLOR_RGB2BGR)
        if len(last_img)==0 or cv2.PSNR(last_img,image)<25:
            rscroll()
            last_img=image
        else:
            vrbs.img_cnt-=1
            os.remove('result/'+str( vrbs.img_cnt )+'_'+name+'.png')
            break
def get_right_infos(image):
    image_canny=cv2.Canny(image,50,100)
    contours=cv2.findContours(image_canny,cv2.CHAIN_APPROX_SIMPLE,cv2.RETR_CCOMP)[0]
    rects=[cv2.cv2.boundingRect(contour) for contour in contours]
    right_rects= filter_rects(rects,vrbs.right_area,vrbs.RNW, vrbs.RXW ,vrbs.RNH , vrbs.RXH)
    right_rects= short_rects(right_rects)
    right_rects.sort(key=lambda x:x[1])
    texts=[]
    infos=[]
    dachengs=[]
    for rect in right_rects:
        left=rect[0]
        up=rect[1]
        right=rect[0]+rect[2]
        down=rect[1]+rect[3]
        mid=left+int( vrbs.RRT/image.shape[1]*(right-left))
        #cv2.imshow('img',image[up:down,left:mid])
        #cv2.waitKey()
        text,info= Image2Result( image[up:down,left:mid])[0:2]
        if text=='达成进度':
            continue 
        results=Image2Result( image[up:down,mid:right])
        dacheng='未达成'
        for result in results:
            if result=='达成':
                dacheng='达成'
        texts.append(text)
        infos.append(info)
        dachengs.append(dacheng)
    return texts,infos,dachengs


def get_right_all_infos():
    workbook= openpyxl.load_workbook('resource/achievements.xlsx')
    worksheet=workbook['Full']
    max_row= worksheet.max_row
    achievements=[]
    all_infos=[]
    for i in range(1,max_row):
        text=worksheet.cell(i+1,2).value
        info=worksheet.cell(i+1,3).value
        if not text in achievements:
            achievements.append(text)
            all_infos.append(info)
        else:
            all_infos[ achievements.index(text)]=info
    all_dacheng=['未达成']*len( achievements)

    workbook= openpyxl.Workbook()
    worksheet=workbook.create_sheet('已完成')
    del workbook['Sheet']
    worksheet_index=1
        

    
    name_list=[]
    text_list=[]
    info_list=[]
    dacheng_list=[]
    all_files=os.listdir('result')
    all_files.sort(key=lambda x:int(x.split('_')[0]))
    for single_file in all_files:
        os.system('cls')
        print(single_file)
        image=cv2.imdecode(np.fromfile(os.path.join('result',single_file),dtype=np.uint8),cv2.IMREAD_COLOR)
        name=single_file.split('.')[0].split('_')[1]
        texts,infos,dachengs=get_right_infos(image)
        for text,info,dacheng in zip(texts,infos,dachengs):
            if not text in text_list:
                print(text,info,dacheng)
                text,info=Find2(text,info, achievements , all_infos)                
                text_list.append(text)
                name_list.append(name)
                info_list.append(info)
                dacheng_list.append(dacheng)
                if dacheng=='达成':
                    all_dacheng[ achievements.index(text)]='达成'
                worksheet.cell(worksheet_index,1).value=name
                worksheet.cell(worksheet_index,2).value=text
                worksheet.cell(worksheet_index,3).value=info
                worksheet.cell(worksheet_index,4).value=dacheng
                workbook.save('成就.xlsx')
                worksheet_index+=1
    worksheet=workbook.create_sheet('未完成')
    worksheet_index=1
    print('最终结果')
    for i,x in enumerate(all_dacheng):
        if x=='未达成':
            print(achievements[i],all_infos[i])
            worksheet.cell(worksheet_index,1).value= achievements[i]
            worksheet.cell(worksheet_index,2).value= all_infos[i]
            worksheet_index+=1
        workbook.save('my_achievements.xlsx')